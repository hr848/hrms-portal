from __future__ import annotations

from io import BytesIO, StringIO
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from app.models.schemas import AttendanceRecord


def records_to_frame(records: list[AttendanceRecord]) -> pd.DataFrame:
    return pd.DataFrame([record.model_dump() for record in records])


def export_csv(records: list[AttendanceRecord]) -> bytes:
    output = StringIO()
    records_to_frame(records).to_csv(output, index=False)
    return output.getvalue().encode("utf-8")


def export_excel(records: list[AttendanceRecord]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        records_to_frame(records).to_excel(writer, index=False, sheet_name="Detailed Records")
    return output.getvalue()


def export_pdf(records: list[AttendanceRecord]) -> bytes:
    output = BytesIO()
    frame = records_to_frame(records).head(100)
    table_data = [list(frame.columns)] + frame.astype(str).values.tolist() if not frame.empty else [["No records uploaded"]]
    document = SimpleDocTemplate(output, pagesize=landscape(letter))
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4ed8")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.grey), ("FONTSIZE", (0, 0), (-1, -1), 6)]))
    document.build([table])
    return output.getvalue()


def export_employee_excel(employee_name: str, records: list[AttendanceRecord]) -> bytes:
    from datetime import date
    from collections import Counter
    from app.rule_engine.attendance_rules import normalize_name
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.series import SeriesLabel

    target_key = normalize_name(employee_name)
    emp_records = [r for r in records if normalize_name(r.employee) == target_key]

    wb = openpyxl.Workbook()

    # ── Sheet 1: Attendance Records ───────────────────────────────────────────
    ws = wb.active
    ws.title = "Attendance Records"
    ws.views.sheetView[0].showGridLines = True

    headers = [
        "Employee", "Date", "Proposed Entry", "Proposed Exit",
        "Actual Entry", "Actual Exit", "Status", "Working Hours", "Raw Punches"
    ]
    ws.append(headers)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    for col_num in range(1, 10):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align if col_num > 1 and col_num < 9 else left_align

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for r in emp_records:
        row_data = [
            r.employee,
            r.date.isoformat() if isinstance(r.date, date) else str(r.date),
            r.proposed_entry,
            r.proposed_exit,
            r.actual_entry or "-",
            r.actual_exit or "-",
            r.status,
            r.working_hours if r.working_hours is not None else "-",
            ", ".join(r.raw_punches) if r.raw_punches else ""
        ]
        ws.append(row_data)
        curr_row = ws.max_row
        for col_num in range(1, 10):
            c = ws.cell(row=curr_row, column=col_num)
            c.border = thin_border
            c.alignment = center_align if col_num > 1 and col_num < 9 else left_align

    ws.append([])
    ws.append([])

    status_counts = Counter(r.status for r in emp_records)

    ws.append(["Attendance Status Summary"])
    summary_title_row = ws.max_row
    ws.merge_cells(start_row=summary_title_row, start_column=1, end_row=summary_title_row, end_column=2)
    title_cell = ws.cell(row=summary_title_row, column=1)
    title_cell.font = Font(name="Calibri", size=12, bold=True)

    ws.append(["Status", "Total Count"])
    summary_header_row = ws.max_row
    for col_num in (1, 2):
        c = ws.cell(row=summary_header_row, column=col_num)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
        c.alignment = left_align

    for status, count in status_counts.items():
        ws.append([status, count])
        curr_row = ws.max_row
        ws.cell(row=curr_row, column=1).alignment = left_align
        ws.cell(row=curr_row, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=curr_row, column=1).border = thin_border
        ws.cell(row=curr_row, column=2).border = thin_border

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.coordinate in ws.merged_cells:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ── Compute chart categories (mirrors frontend logic) ─────────────────────
    present = 0
    on_time_entry = 0
    on_time_exit = 0
    late_entry = 0
    entry_missing = 0
    early_exit = 0
    holiday = 0
    wfh = 0
    exit_missing = 0
    leave = 0
    sunday = 0

    valid_hours = []
    for r in emp_records:
        s = r.status.lower()
        has_punch = bool(r.actual_entry or r.actual_exit)
        is_not_absent = 'no attendance' not in s and 'leave' not in s and 'holiday' not in s and 'sunday' not in s
        if (has_punch and is_not_absent) or s == 'present':
            present += 1
        if 'ontime entry' in s or 'ontime entry & exit' in s:
            on_time_entry += 1
        if 'ontime exit' in s or 'ontime entry & exit' in s:
            on_time_exit += 1
        if 'late entry' in s:
            late_entry += 1
        if 'entry missing' in s:
            entry_missing += 1
        if 'early exit' in s:
            early_exit += 1
        if 'holiday' in s:
            holiday += 1
        if 'wfh' in s:
            wfh += 1
        if 'exit missing' in s:
            exit_missing += 1
        if 'leave' in s:
            leave += 1
        if 'sunday' in s:
            sunday += 1
        if r.actual_entry and r.actual_exit and r.working_hours is not None:
            valid_hours.append(r.working_hours)

    avg_hours = round(sum(valid_hours) / len(valid_hours), 2) if valid_hours else 0

    raw_categories = [
        ("PRESENT DAY",    present),
        ("ON TIME ENTRY",  on_time_entry),
        ("ON TIME EXIT",   on_time_exit),
        ("LATE ENTRY",     late_entry),
        ("ENTRY MISSING",  entry_missing),
        ("EXIT MISSING",   exit_missing),
        ("EARLY EXIT",     early_exit),
        ("HOLIDAY",        holiday),
        ("WFH",            wfh),
        ("LEAVE",          leave),
        ("SUNDAY",         sunday),
    ]
    # Sort descending by count
    categories = sorted(raw_categories, key=lambda x: x[1], reverse=True)

    # ── Place Chart and Stats Table on the same sheet (Columns K & L) ─────────
    # Write Average Working Hours badge in merged K1:M1
    avg_label = f"Avg Working Hours: {avg_hours}h  {'✓ Above 9h' if avg_hours >= 9 else '✗ Below 9h'}"
    ws.merge_cells("K1:M1")
    avg_cell = ws.cell(row=1, column=11, value=avg_label)
    avg_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    avg_cell.fill = PatternFill(
        start_color="00B050" if avg_hours >= 9 else "FF0000",
        end_color="00B050" if avg_hours >= 9 else "FF0000",
        fill_type="solid"
    )
    avg_cell.alignment = center_align

    # Chart data table headers
    ws.cell(row=2, column=11, value="Category").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=2, column=12, value="Count").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=2, column=11).border = thin_border
    ws.cell(row=2, column=12).border = thin_border
    ws.cell(row=2, column=11).alignment = left_align
    ws.cell(row=2, column=12).alignment = center_align

    # Excel's default chart palette colors (Office theme)
    EXCEL_THEME_COLORS = [
        "4F81BD",  # 1. Steel Blue
        "C0504D",  # 2. Burgundy Red
        "9BBB59",  # 3. Olive Green
        "8064A2",  # 4. Purple
        "4BABC6",  # 5. Teal/Cyan
        "F79646",  # 6. Orange
        "8FA9C4",  # 7. Light Steel Blue
        "A9A9A9",  # 8. Dark Grey
        "D3D3D3",  # 9. Light Grey
        "C0C0C0",  # 10. Silver
        "708090",  # 11. Slate Grey
    ]

    # Fill data
    for i, (name, count) in enumerate(categories):
        r_idx = 3 + i
        hex_color = EXCEL_THEME_COLORS[i % len(EXCEL_THEME_COLORS)]
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        
        c1 = ws.cell(row=r_idx, column=11, value=name)
        c2 = ws.cell(row=r_idx, column=12, value=count)
        
        c1.border = thin_border
        c2.border = thin_border
        c1.fill = fill
        c2.fill = fill
        c1.font = font
        c2.font = font
        
        c1.alignment = left_align
        c2.alignment = Alignment(horizontal="right", vertical="center")

    ws.column_dimensions["K"].width = 18
    ws.column_dimensions["L"].width = 10

    # Add native Excel Bar Chart
    data_ref = Reference(ws, min_col=12, min_row=2, max_row=2 + len(categories))
    cats_ref = Reference(ws, min_col=11, min_row=3, max_row=2 + len(categories))

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Attendance & Time Tracking Overview"
    chart.y_axis.title = "Count / Days"
    chart.legend = None
    chart.width = 18
    chart.height = 12

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Enable data labels (shows figures above bars)
    s = chart.series[0]
    s.dLbls = DataLabelList()
    s.dLbls.showVal = True
    s.dLbls.showLegendKey = False
    s.dLbls.showCatName = False
    s.dLbls.showSerName = False

    # Add chart at column N, row 3
    ws.add_chart(chart, "N3")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
