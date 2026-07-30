"""
excel_parser.py
===============
Parses two Excel file formats for the Attendance Analytics Portal.

Proposed Time Sheet
-------------------
Standard tabular format:  one row per employee, columns for name / dept /
proposed-entry / proposed-exit.  Flexible header detection handles many
real-world column naming variations.

Attendance Log
--------------
Supports TWO formats automatically:

  Format A  –  Standard (row = employee, column = date)
      • Column headers are actual date values (date objects or date strings).
      • One cell per employee per day.

  Format B  –  Day-Number (the format produced by most biometric exporters)
      • Title row contains the month range, e.g.
            "Att. Time 2026-06-01 ~ 2026-06-30"
      • A header row holds day numbers:  1  2  3 … 30
      • Each employee occupies 1 OR 2 Excel rows:
            Row A – contains "Name: <employee>" in a (merged) cell
                    PLUS the first punch of each day in the day-columns.
            Row B – (optional) contains additional / exit punches for each day.
      • Cells may hold multiple time strings separated by newlines.

Attendance Parsing Rules (Rule Engine contract)
-----------------------------------------------
  Rule 1 : times  < 15:00  → Entry candidates   (keep FIRST)
  Rule 2 : times ≥ 15:00  → Exit  candidates   (keep LAST)
  Rule 3 : blank / no punches → No Attendance
  Rule 4 : entry only          → Exit Missing
  Rule 5 : exit  only          → Entry Missing
  Rule 6 : all remaining times stored as raw_punches for display
"""

from __future__ import annotations

import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from app.models.schemas import AttendanceRecord, ProposedShift
from app.rule_engine.attendance_rules import (
    AttendanceRuleEngine,
    format_time,
    minutes_between,
    normalize_name,
    parse_time_value,
)

# ── Column-name aliases for proposed sheet detection ──────────────────────────
NAME_ALIASES = [
    "employee name", "employee", "name", "staff name", "emp name",
    "employee full name", "full name",
]
DEPT_ALIASES = ["department", "dept", "division", "team", "function"]
ENTRY_ALIASES = [
    "proposed entry time", "proposed entry", "entry time", "shift entry",
    "in time", "intime", "in-time", "office in time", "office entry time",
    "proposed in time", "proposed office in time", "start time",
    "shift start", "login time", "reporting time", "from time", "entry", "in",
]
EXIT_ALIASES = [
    "proposed exit time", "proposed exit", "exit time", "shift exit",
    "out time", "outtime", "out-time", "office out time", "office exit time",
    "proposed out time", "proposed office out time", "end time",
    "shift end", "logout time", "leaving time", "to time", "exit", "out",
]
SHIFT_ALIASES = [
    "proposed office timing", "proposed timing", "office timing",
    "shift timing", "timing", "shift", "working hours",
]

# Regex: marks a cell as metadata (not attendance data) if matched
_METADATA_RE = re.compile(r"\b(Name|Dept|ID)\s*:", re.IGNORECASE)
# Regex: extract employee name after "Name:"
_NAME_RE = re.compile(
    r"\bName\s*:\s*(.+?)(?=\s{2,}|\t|\bDept\b|$)", re.IGNORECASE
)
# Regex: extract YYYY-MM or YYYY/MM from any string
_YEARMONTH_RE = re.compile(r"(\d{4})[/-](\d{1,2})[/-]\d{1,2}")


# ── Utility helpers ────────────────────────────────────────────────────────────

def _clean_header(value: Any) -> str:
    text = normalize_name(value)
    for token in ["_", "-", "/", "\n", "\r"]:
        text = text.replace(token, " ")
    return " ".join(text.split())


def _find_column(columns: list[Any], aliases: list[str]) -> Any | None:
    normalized = {_clean_header(col): col for col in columns}
    compact = {k.replace(" ", ""): v for k, v in normalized.items()}
    for alias in aliases:
        ca = _clean_header(alias)
        if ca in normalized:
            return normalized[ca]
        if ca.replace(" ", "") in compact:
            return compact[ca.replace(" ", "")]
    for nk, original in normalized.items():
        if any(_clean_header(a) in nk for a in aliases):
            return original
    return None


def _parse_date_header(header: Any) -> date | None:
    if isinstance(header, datetime):
        return header.date()
    if isinstance(header, date):
        return header
    parsed = pd.to_datetime(header, errors="coerce", dayfirst=True)
    return None if pd.isna(parsed) else parsed.date()


def _extract_times(value: Any) -> list[time]:
    matches = re.findall(r"(?:[01]?\d|2[0-3])\s*[:.]?\s*[0-5]\d", str(value or ""))
    return [t for m in matches if (t := parse_time_value(m)) is not None]


def _read_excel_with_detected_header(path: Path) -> pd.DataFrame:
    """Load an Excel file, auto-detecting which row is the true header."""
    sheets = pd.read_excel(path, header=None, sheet_name=None)
    best_sheet, best_row, best_score = None, 0, -1
    for _, raw in sheets.items():
        if raw.dropna(how="all").empty:
            continue
        for row_index in range(min(20, len(raw))):
            values = list(raw.iloc[row_index].fillna(""))
            score = 0
            if _find_column(values, NAME_ALIASES):
                score += 3
            if _find_column(values, ENTRY_ALIASES):
                score += 2
            if _find_column(values, EXIT_ALIASES):
                score += 2
            if _find_column(values, SHIFT_ALIASES):
                score += 2
            score += 2 if sum(1 for v in values if _parse_date_header(v)) >= 2 else 0
            if score > best_score:
                best_sheet, best_row, best_score = raw, row_index, score
    if best_sheet is None:
        return pd.read_excel(path)
    if best_score <= 0:
        return best_sheet.dropna(how="all")
    cols = [
        str(v).strip() if str(v).strip() and str(v).lower() != "nan"
        else f"Column {i + 1}"
        for i, v in enumerate(best_sheet.iloc[best_row])
    ]
    frame = best_sheet.iloc[best_row + 1:].copy()
    frame.columns = cols
    return frame.dropna(how="all")


def _cell_str(value: Any) -> str:
    """Return a clean string from a raw cell value, or '' if empty/None."""
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s.lower() in ("none", "nan", "") else s


# ── Main parser class ─────────────────────────────────────────────────────────

class AttendanceExcelParser:
    """
    Parses both the Proposed Time Sheet and Attendance Log Excel files.
    Delegates status computation to AttendanceRuleEngine.
    """

    def __init__(self, rule_engine: AttendanceRuleEngine | None = None) -> None:
        self.rule_engine = rule_engine or AttendanceRuleEngine()

    # ── Public API ────────────────────────────────────────────────────────────

    def parse_proposed_sheet(
        self, path: Path
    ) -> tuple[dict[str, ProposedShift], list[str]]:
        """Return {normalised_name: ProposedShift} and a list of warnings."""
        frame = _read_excel_with_detected_header(path)
        warnings: list[str] = []
        columns = list(frame.columns)

        name_col  = _find_column(columns, NAME_ALIASES)
        dept_col  = _find_column(columns, DEPT_ALIASES)
        entry_col = _find_column(columns, ENTRY_ALIASES)
        exit_col  = _find_column(columns, EXIT_ALIASES)
        shift_col = _find_column(columns, SHIFT_ALIASES)

        if name_col is None:
            name_col = columns[0] if columns else None
        if name_col is None:
            raise ValueError("Proposed sheet missing columns: Employee Name")
        if entry_col is None and shift_col is None:
            raise ValueError(
                "Proposed sheet missing columns: "
                "Proposed Entry Time or Proposed Office Timing"
            )
        if exit_col is None and shift_col is None:
            raise ValueError(
                "Proposed sheet missing columns: "
                "Proposed Exit Time or Proposed Office Timing"
            )

        shifts: dict[str, ProposedShift] = {}
        for _, row in frame.iterrows():
            name = str(row.get(name_col, "")).strip()
            if not name or name.lower() == "nan":
                continue
            entry     = parse_time_value(row.get(entry_col)) if entry_col else None
            exit_time = parse_time_value(row.get(exit_col))  if exit_col  else None
            if (not entry or not exit_time) and shift_col:
                ts = _extract_times(row.get(shift_col))
                if not entry and ts:
                    entry = ts[0]
                if not exit_time and len(ts) >= 2:
                    exit_time = ts[-1]
            if not entry or not exit_time:
                warnings.append(f"Skipping {name}: invalid proposed shift timing")
                continue
            shifts[normalize_name(name)] = ProposedShift(
                employee_name=name,
                department=(
                    ""
                    if dept_col is None or pd.isna(row.get(dept_col))
                    else str(row.get(dept_col)).strip()
                ),
                proposed_entry=format_time(entry) or "",
                proposed_exit=format_time(exit_time) or "",
            )

        if not shifts:
            raise ValueError(
                "No valid employees found in Proposed Office Time Sheet. "
                "Check employee name and timing columns."
            )
        return shifts, warnings

    def parse_attendance(
        self,
        proposed: dict[str, ProposedShift],
        path: Path,
    ) -> tuple[list[AttendanceRecord], list[str]]:
        """
        Auto-detect format and parse the attendance log.

        Tries Format A (standard date-column layout) first.
        Falls back to Format B (day-number columns with inline employee names)
        if no date-typed column headers are found.
        """
        # ── Try Format A (standard: rows = employees, cols = dates) ──────
        try:
            frame   = _read_excel_with_detected_header(path)
            columns = list(frame.columns)
            name_col = _find_column(columns, NAME_ALIASES) or columns[0]
            date_cols = [
                (col, d)
                for col in columns
                if col != name_col and (d := _parse_date_header(col))
            ]
            # Validate that Format A dates are genuine:
            # - Must have at least 2 distinct date values (not all the same day)
            # - Dates must not all map to today (pandas coerces time strings like
            #   "11:28\n19:02" to today's date, producing a false positive)
            if date_cols:
                from datetime import date as _date
                unique_dates = {d for _, d in date_cols}
                today = _date.today()
                genuine = len(unique_dates) >= 2 and not all(d == today for _, d in date_cols)
                if genuine:
                    return self._parse_format_a(proposed, frame, name_col, date_cols)
        except Exception:
            pass  # fall through to Format B

        # ── Fall back to Format B (day-number columns) ────────────────────
        return self._parse_format_b(path, proposed)

    # ── Format A: standard rows=employees, cols=dates ─────────────────────────

    def _parse_format_a(
        self,
        proposed: dict[str, ProposedShift],
        frame: pd.DataFrame,
        name_col: Any,
        date_cols: list[tuple[Any, date]],
    ) -> tuple[list[AttendanceRecord], list[str]]:
        warnings: list[str] = []
        records:  list[AttendanceRecord] = []

        by_name = {
            normalize_name(row.get(name_col)): row
            for _, row in frame.iterrows()
            if normalize_name(row.get(name_col))
        }

        for key, shift in proposed.items():
            row = by_name.get(key)
            prop_entry = parse_time_value(shift.proposed_entry)
            prop_exit  = parse_time_value(shift.proposed_exit)
            if row is not None and prop_entry and prop_exit:
                for col, rec_date in date_cols:
                    parsed = self.rule_engine.parse_cell(row.get(col))
                    status = self.rule_engine.determine_status(
                        prop_entry, prop_exit,
                        parsed.entry, parsed.exit, parsed.invalid,
                    )
                    records.append(
                        self._build_record(
                            shift, rec_date, prop_entry, prop_exit, parsed, status
                        )
                    )
            else:
                warnings.append(f"Attendance Not Found for {shift.employee_name}")
                for _, rec_date in date_cols:
                    status_val = "Attendance Not Found"
                    if isinstance(rec_date, date) and rec_date.weekday() == 6:
                        status_val = "sunday"
                    records.append(AttendanceRecord(
                        employee=shift.employee_name,
                        department=shift.department,
                        date=rec_date,
                        proposed_entry=shift.proposed_entry,
                        proposed_exit=shift.proposed_exit,
                        status=status_val,
                        notes="Employee exists in Proposed Sheet but not Attendance Log",
                    ))
        return records, warnings

    # ── Format B: day-number column headers with inline "Name:" ───────────────

    def _parse_format_b(
        self,
        path: Path,
        proposed: dict[str, ProposedShift],
    ) -> tuple[list[AttendanceRecord], list[str]]:
        """
        Parse the biometric-exporter format:

          Row 1-3 : Title / metadata  →  e.g. "Att. Time 2026-06-01 ~ 2026-06-30"
          Row 4   : Day-number headers:  1  2  3 … 30
          Row 5+  : Employee blocks (1 or 2 rows per employee):
                     • Row A contains "Name: <employee>" in a merged cell AND
                       the first punch for each day in the corresponding day-column.
                     • Row B (optional, if it does NOT start a new employee)
                       contains additional / exit punches per day.

        Cell values may be:
          • A datetime.time  object (single punch, stored as Excel time fraction)
          • A plain string   "HH:MM"
          • A multi-line string "HH:MM\nHH:MM"  (multiple punches; Alt+Enter)
          • None / blank

        All cell values for a given employee+day (rows A + B) are concatenated
        with newlines and fed into the existing rule_engine.parse_cell(), which
        applies all 8 parsing rules from the project specification.
        """
        warnings: list[str] = []
        records:  list[AttendanceRecord] = []

        wb = openpyxl.load_workbook(path, data_only=True)

        for ws in wb.worksheets:

            # ── Step 1: Extract year + month from title rows ───────────────
            year_month: tuple[int, int] | None = None
            for row in ws.iter_rows(max_row=10):
                for cell in row:
                    if cell.value is None:
                        continue
                    m = _YEARMONTH_RE.search(str(cell.value))
                    if m:
                        year_month = int(m.group(1)), int(m.group(2))
                        break
                if year_month:
                    break

            if year_month is None:
                continue  # try next sheet

            year, month = year_month

            # ── Step 2: Find the row that holds day-numbers 1 … 31 ────────
            #
            # We look for the first row where ≥ 20 cells contain integers
            # in the range [1, 31].  This avoids false positives from title
            # rows that might contain small numbers.
            #
            day_col_map: dict[int, date] = {}   # excel_column_index → date
            header_row_idx: int = 0

            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                day_nums: dict[int, int] = {}
                for cell in row:
                    if cell.value is None:
                        continue
                    raw = str(cell.value).strip()
                    try:
                        n = int(float(raw))
                        if 1 <= n <= 31:
                            day_nums[cell.column] = n
                    except (ValueError, TypeError):
                        pass

                if len(day_nums) >= 20:
                    # ── Validate contiguous sequential run ────────────────────
                    # Day numbers must be sequential (1, 2, 3 …) AND their
                    # Excel columns must be contiguous (no gaps between them).
                    # This prevents stray row-number labels from being treated
                    # as day-header columns.
                    sorted_by_col = sorted(day_nums.items())          # [(col, day), …]
                    # Find the longest contiguous sequential block
                    best_block: list[tuple[int, int]] = []
                    current_block: list[tuple[int, int]] = [sorted_by_col[0]]
                    for j in range(1, len(sorted_by_col)):
                        prev_col, prev_day = sorted_by_col[j - 1]
                        this_col, this_day = sorted_by_col[j]
                        if this_col == prev_col + 1 and this_day == prev_day + 1:
                            current_block.append(sorted_by_col[j])
                        else:
                            if len(current_block) > len(best_block):
                                best_block = current_block
                            current_block = [sorted_by_col[j]]
                    if len(current_block) > len(best_block):
                        best_block = current_block

                    if len(best_block) < 20:
                        continue           # not a valid day-number header row

                    header_row_idx = row_idx
                    for col, d in best_block:
                        try:
                            day_col_map[col] = date(year, month, d)
                        except ValueError:
                            pass  # skip invalid dates (e.g. Feb 30)
                    break

            if not day_col_map:
                continue  # this sheet has no day-number header, try next

            # ── Step 3: Load all data rows into memory ─────────────────────
            all_rows = list(ws.iter_rows())

            # ── Step 4: Find employee rows ("Name:" marker) ────────────────
            #
            # An "employee row" is any row below the day-header that contains
            # a cell whose string value matches /Name\s*:/.
            # We collect their 0-based indices for later processing.
            #
            emp_row_indices: list[int] = []
            for i in range(header_row_idx, len(all_rows)):
                for cell in all_rows[i]:
                    if cell.value and re.search(
                        r"\bName\s*:", str(cell.value), re.IGNORECASE
                    ):
                        emp_row_indices.append(i)
                        break

            if not emp_row_indices:
                continue  # no employees found on this sheet

            # ── Step 5: Build attendance records ───────────────────────────
            for emp_i in emp_row_indices:
                emp_row = all_rows[emp_i]

                # Extract employee name from the "Name: <name>" cell or adjacent cell
                employee_name: str | None = None
                for i_cell, cell in enumerate(emp_row):
                    if cell.value is None:
                        continue
                    val_str = str(cell.value).strip()
                    m = _NAME_RE.search(val_str)
                    if m:
                        candidate = m.group(1).strip()
                        if len(candidate) > 1:
                            employee_name = candidate
                            break
                    elif re.search(r"^\s*Name\s*:\s*$", val_str, re.IGNORECASE):
                        # Look ahead in the next few cells for the actual name
                        for next_cell in emp_row[i_cell+1:i_cell+5]:
                            if next_cell.value:
                                candidate = str(next_cell.value).strip()
                                if len(candidate) > 1 and not candidate.lower().startswith("dept"):
                                    employee_name = candidate
                                    break
                        if employee_name:
                            break

                if not employee_name:
                    continue

                name_key = normalize_name(employee_name)
                shift    = proposed.get(name_key)

                if shift is None:
                    warnings.append(
                        f"Attendance Not Found for {employee_name} "
                        f"(not in Proposed Time Sheet)"
                    )
                    continue

                prop_entry = parse_time_value(shift.proposed_entry)
                prop_exit  = parse_time_value(shift.proposed_exit)
                if not prop_entry or not prop_exit:
                    warnings.append(
                        f"Skipping {employee_name}: invalid proposed shift timing"
                    )
                    continue

                # Determine if the NEXT row belongs to this employee (not a new one).
                # In the 2-row format, the second row holds additional punches.
                next_i         = emp_i + 1
                next_has_name  = False
                if next_i < len(all_rows):
                    for cell in all_rows[next_i]:
                        if cell.value and re.search(
                            r"\bName\s*:", str(cell.value), re.IGNORECASE
                        ):
                            next_has_name = True
                            break

                # Build column → raw-value maps for current row and (optionally) next row
                cur_vals: dict[int, Any] = {
                    cell.column: cell.value for cell in emp_row
                }
                next_vals: dict[int, Any] = (
                    {cell.column: cell.value for cell in all_rows[next_i]}
                    if next_i < len(all_rows) and not next_has_name
                    else {}
                )

                # ── Step 6: Process each day column ────────────────────────
                for col, rec_date in day_col_map.items():
                    # Combine punches from both rows into one multi-line string.
                    # Skip cells that are clearly metadata (merged Name/Dept cells).
                    parts: list[str] = []
                    for raw_val in (cur_vals.get(col), next_vals.get(col)):
                        s = _cell_str(raw_val)
                        if not s:
                            continue
                        if _METADATA_RE.search(s):
                            # This day-column is occupied by a merged Name/Dept cell
                            continue
                        # Handle openpyxl returning datetime.time objects
                        if isinstance(raw_val, time):
                            s = format_time(raw_val) or s
                        parts.append(s)

                    combined = "\n".join(parts) if parts else None

                    # Apply the 8-rule attendance parsing (Rule Engine)
                    parsed = self.rule_engine.parse_cell(combined)
                    status = self.rule_engine.determine_status(
                        prop_entry, prop_exit,
                        parsed.entry, parsed.exit, parsed.invalid,
                    )
                    records.append(
                        self._build_record(
                            shift, rec_date,
                            prop_entry, prop_exit,
                            parsed, status,
                        )
                    )

            # If we found records on this sheet, stop processing more sheets
            if records:
                break

        if not records and not warnings:
            raise ValueError(
                "Attendance log format not recognised. "
                "Expected either (a) date-column headers, or "
                "(b) a title row with 'YYYY-MM-DD' and a day-number header row (1…30)."
            )

        return records, warnings

    # ── Shared record builder ─────────────────────────────────────────────────

    def _build_record(
        self,
        shift: ProposedShift,
        rec_date: date,
        prop_entry: Any,
        prop_exit: Any,
        parsed: Any,
        status: str,
    ) -> AttendanceRecord:
        if isinstance(rec_date, date) and rec_date.weekday() == 6:
            if not parsed.entry and not parsed.exit:
                status = "sunday"
                
        working_hours  = (
            round(minutes_between(parsed.entry, parsed.exit) / 60, 2)
            if parsed.entry and parsed.exit else None
        )
        arrival_delay  = (
            max(0, minutes_between(prop_entry, parsed.entry))
            if parsed.entry else None
        )
        early_exit_min = (
            max(0, minutes_between(parsed.exit, prop_exit))
            if parsed.exit else None
        )
        overtime_min   = (
            max(0, minutes_between(parsed.entry, parsed.exit) - 9 * 60)
            if parsed.entry and parsed.exit else None
        )
        return AttendanceRecord(
            employee=shift.employee_name,
            department=shift.department,
            date=rec_date,
            proposed_entry=shift.proposed_entry,
            proposed_exit=shift.proposed_exit,
            actual_entry=format_time(parsed.entry),
            actual_exit=format_time(parsed.exit),
            status=status,
            working_hours=working_hours,
            arrival_delay_minutes=arrival_delay,
            early_exit_minutes=early_exit_min,
            overtime_minutes=overtime_min,
            raw_punches=parsed.raw_punches,
        )
